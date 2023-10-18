import json
import random
import re
import openpyxl


# FUNCTIONS FOR WORKING WITH JSON FILES


def read_json(file):
    with open(file, encoding="UTF-8") as json_file:
        dataset = json.load(json_file)
    return dataset


def fetch_quote_json(index_of_quote, dataset):
    quote = dataset[index_of_quote]['Quote'].strip()
    author = re.sub(' {2}', ' ', string=dataset[index_of_quote]['Author'])
    category = dataset[index_of_quote]['Category'].capitalize()
    return quote, author, category


def fetch_random_quote_json(dataset):
    random_quote_number = random.randint(0, len(dataset) - 1)
    random_quote = fetch_quote_json(random_quote_number, dataset)
    return random_quote


def get_unique_attributes_of_json(dataset, attribute):
    attributes = []
    for data in dataset:
        current_attribute = data[attribute]
        if current_attribute == '':
            current_attribute = "uncategorized"
        if current_attribute not in attributes:
            attributes.append(current_attribute)
    return attributes


# FUNCTIONS FOR WORKING WITH EXCEL FILESpipenv she

def read_excel(file):
    wb = openpyxl.load_workbook(file)
    dataset = wb['quotes']
    return dataset


def fetch_quote_excel(index_of_quote, dataset):
    index_of_quote += 1
    quote = (dataset.cell(index_of_quote, 1)).value
    author = (dataset.cell(index_of_quote, 2)).value
    category = (dataset.cell(index_of_quote, 3)).value
    return index_of_quote, quote, author, category


def fetch_random_quote_excel(dataset):
    random_quote_number = random.randint(1, dataset.max_row - 1)
    random_quote = fetch_quote_excel(random_quote_number, dataset)
    return random_quote


def get_unique_attributes_of_excel(dataset, attribute):
    column = 3
    if attribute == 'Category':
        column = 3
    if attribute == 'Author':
        column = 2
    unique_categories = []
    for row in range(2, dataset.max_row - 1):
        category = (dataset.cell(row, column)).value
        if category not in unique_categories:
            unique_categories.append(category)
    return unique_categories
